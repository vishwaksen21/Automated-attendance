"""
Export Module for Attendance System
Handles exporting attendance data to Excel and PDF formats
"""

import os
import pandas as pd
from datetime import datetime
import logger_config

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

logger = logger_config.get_logger('Export')


def export_to_excel(csv_path, output_path=None):
    """
    Export attendance CSV to formatted Excel file
    
    Args:
        csv_path (str): Path to the CSV file
        output_path (str): Optional output path for Excel file
    
    Returns:
        str: Path to created Excel file or None if failed
    """
    if not EXCEL_AVAILABLE:
        logger.error("Excel export failed - openpyxl not installed")
        return None
    
    try:
        # Read CSV
        df = pd.read_csv(csv_path)
        
        # Generate output path if not provided
        if output_path is None:
            base_name = os.path.splitext(csv_path)[0]
            output_path = base_name + '.xlsx'
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Attendance')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            # Style the header row
            header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Style data cells
            data_alignment = Alignment(horizontal="left", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = data_alignment
                    cell.border = border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add metadata
            worksheet.sheet_properties.tabColor = "0078D7"
        
        logger.info(f"Excel export successful: {output_path}")
        return output_path
        
    except Exception as e:
        logger_config.log_error('Export', f'Excel export failed for {csv_path}', e)
        return None


def export_to_pdf(csv_path, output_path=None, title=None):
    """
    Export attendance CSV to PDF file
    
    Args:
        csv_path (str): Path to the CSV file
        output_path (str): Optional output path for PDF file
        title (str): Optional title for the PDF
    
    Returns:
        str: Path to created PDF file or None if failed
    """
    if not PDF_AVAILABLE:
        logger.error("PDF export failed - reportlab not installed")
        return None
    
    try:
        # Read CSV
        df = pd.read_csv(csv_path)
        
        # Generate output path if not provided
        if output_path is None:
            base_name = os.path.splitext(csv_path)[0]
            output_path = base_name + '.pdf'
        
        # Generate title if not provided
        if title is None:
            filename = os.path.basename(csv_path)
            title = f"Attendance Report - {filename}"
        
        # Create PDF
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#0078D7'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Add generation info
        info_style = styles['Normal']
        info_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Records: {len(df)}"
        elements.append(Paragraph(info_text, info_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Prepare table data
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table
        table = Table(table_data)
        
        # Style table
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078D7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        logger.info(f"PDF export successful: {output_path}")
        return output_path
        
    except Exception as e:
        logger_config.log_error('Export', f'PDF export failed for {csv_path}', e)
        return None


def export_attendance(csv_path, format='both', output_dir=None):
    """
    Export attendance to specified format(s)
    
    Args:
        csv_path (str): Path to the CSV file
        format (str): 'excel', 'pdf', or 'both'
        output_dir (str): Optional output directory
    
    Returns:
        dict: Dictionary with 'excel' and 'pdf' keys containing file paths or None
    """
    results = {'excel': None, 'pdf': None}
    
    if not os.path.exists(csv_path):
        logger.error(f"Export failed - CSV file not found: {csv_path}")
        return results
    
    # Determine output paths
    base_name = os.path.basename(csv_path)
    name_without_ext = os.path.splitext(base_name)[0]
    
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    if format in ['excel', 'both']:
        excel_path = os.path.join(output_dir, name_without_ext + '.xlsx') if output_dir else None
        results['excel'] = export_to_excel(csv_path, excel_path)
    
    if format in ['pdf', 'both']:
        pdf_path = os.path.join(output_dir, name_without_ext + '.pdf') if output_dir else None
        results['pdf'] = export_to_pdf(csv_path, pdf_path)
    
    return results


# Check if required packages are installed
def check_export_capabilities():
    """
    Check which export formats are available
    
    Returns:
        dict: Dictionary with boolean values for each format
    """
    return {
        'excel': EXCEL_AVAILABLE,
        'pdf': PDF_AVAILABLE
    }
